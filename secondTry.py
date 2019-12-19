import telebot
import os
import time
import random
import openpyxl


bot = telebot.TeleBot('947227617:AAEBlhomRlRuqeTymeM-zVo9Pn6JmxwgVf4')

workbook = openpyxl.load_workbook(filename="/Users/elinaaptineeva/Desktop/Music.xlsx")
ny_sheet = workbook["NewYear_id"]
love_sheet = workbook["InLove_id"]
gym_sheet = workbook["Gym_id"]
sad_sheet = workbook["Sad_id"]
party_sheet = workbook["Party_id"]
nostalgia_sheet = workbook["Nostalgia_id"]
studying_sheet = workbook["Studying_id"]
pop_sheet = workbook["Pop_id"]
alternative_sheet = workbook["Alternative_id"]
rock_sheet = workbook["Rock_id"]
randb_sheet = workbook["RandB_id"]
electro_sheet = workbook["Electro_id"]
sleep_sheet = workbook["ForSleep_id"]
acoustics_sheet = workbook["Acoustics_id"]

ny_list = [ny_sheet.cell(row=i,column=1).value for i in range(1,13)]
love_list = [love_sheet.cell(row=i,column=1).value for i in range(1,21)]
gym_list = [gym_sheet.cell(row=i,column=1).value for i in range(1,21)]
sad_list = [sad_sheet.cell(row=i,column=1).value for i in range(1,21)]
party_list = [party_sheet.cell(row=i,column=1).value for i in range(1,21)]
nostalgia_list = [nostalgia_sheet.cell(row=i,column=1).value for i in range(1,21)]
studying_list = [studying_sheet.cell(row=i,column=1).value for i in range(1,21)]
pop_list = [pop_sheet.cell(row=i,column=1).value for i in range(1,21)]
alternative_list = [alternative_sheet.cell(row=i,column=1).value for i in range(1,21)]
rock_list = [rock_sheet.cell(row=i,column=1).value for i in range(1,21)]
randb_list = [randb_sheet.cell(row=i,column=1).value for i in range(1,21)]
electro_list = [electro_sheet.cell(row=i,column=1).value for i in range(1,21)]
sleep_list = [sleep_sheet.cell(row=i,column=1).value for i in range(1,21)]
acoustics_list = [acoustics_sheet.cell(row=i,column=1).value for i in range(1,21)]
#birthday_list = рандом из всех листов
#top просто 5 наших песен




#keyboard = telebot.types.ReplyKeyboardMarkup()
#keyboard.row('Привет', 'Пока', 'ты лох?', 'песенка', 'музыка')

keyboard_mood = telebot.types.ReplyKeyboardMarkup()
keyboard_mood.row('New Year', 'In love', 'Gym')
keyboard_mood.row('Party', 'Nostalgia', 'Studying')
keyboard_mood.row('Poetry', 'Sad', 'Назад')

keyboard_menu = telebot.types.ReplyKeyboardMarkup()
keyboard_menu.row('По настроению', 'По жанрам')
keyboard_menu.row('По дате рождения', 'Топ песен')
keyboard_menu.row('Для сна', 'Акустика')

keyboard_genre = telebot.types.ReplyKeyboardMarkup()
keyboard_genre.row('Pop', 'Альтернатива')
keyboard_genre.row('Рок', 'R&B', 'Электроника')
keyboard_genre.row('Назад')

#Списки file_id аудио для настроения

NewYear_id = ['CQADAgADfgUAAkn8yEuQTb7jsefMwRYE', 'CQADAgADuQMAAlB-0Esk5DfJL3PL3xYE', 'CQADAgADugMAAlB-0EsKKBL745odthYE',
             'CQADAgADiAUAAlHpyEtHo19BOckZBhYE', 'CQADAgADigUAAlHpyEsncQwqd_wN-hYE', 'CQADAgADEgUAAuKn0Evq-j1RUa4aERYE',
             'CQADAgADDgYAAkn80EtWBclWfG6KhhYE', 'CQADAgADDwYAAkn80Evvwca7YbFaoBYE', 'CQADAgADEAYAAkn80Euh1lFcu-MJNBYE',
             'CQADAgADzwYAAkn80Etwrhhg48UHVhYE', 'CQADAgAD0AYAAkn80EsKVrpSlJGvPhYE', 'CQADAgADWwQAAuKn2Evvw1mMqBEBZBYE']

InLove_id = ['CQADAgADJgQAAlB-0Es96dEgDKGQghYE', 'CQADAgADGQYAAkn80EvnNZ9FNewtwxYE', 'CQADAgADGwYAAkn80EuN3DNk4QMXhxYE',
          'CQADAgADFgUAAuKn0Etzpel0EkiqURYE', 'CQADAgADGgUAAuKn0EsBD2q95fkLaRYE', 'CQADAgADIwYAAkn80Ev1DMxu78JQXhYE',
          'CQADAgADJAYAAkn80EsnID9m4-LQDRYE', 'CQADAgADJgYAAkn80EuDlBByM1LzphYE', 'CQADAgADJwYAAkn80Et4CCNhik8t4RYE',
          'CQADAgADKAYAAkn80Eshm3ICw0V4kxYE', 'CQADAgADHwUAAuKn0EvxUqwOz4SkVRYE', 'CQADAgADKQYAAkn80EtATnQVRpPWhxYE',
          'CQADAgADKgYAAkn80EuSp890BjtcnhYE', 'CQADAgADKwYAAkn80Ev_cdb_6Ylm4RYE', 'CQADAgADLAYAAkn80EtAhTYEUxHk0xYE',
          'CQADAgADJAUAAuKn0EsOjFP7zHSMGxYE', 'CQADAgADJQUAAuKn0EvNjca2H-iMVxYE', 'CQADAgADLQYAAkn80EsmWwg4rNxR8xYE',
          'CQADAgADJwUAAuKn0EuH_xVjSLMGERYE', 'CQADAgADLwYAAkn80EtQPbwwElckAAEWBA', 'CQADAgADKQQAAlB-0Etro6O-CxIXqhYE']

Gym_id = ['CQADAgADaAQAAuKn2Ev6bqsxt0Z9LhYE', 'CQADAgADaQQAAuKn2EvKeDre7P00wRYE', 'CQADAgADagQAAuKn2EsNZKbj2w55FxYE',
          'CQADAgAD2QYAAkn80EvOWq-vhPwrnRYE', 'CQADAgADbgQAAuKn2EspMN9EhoWJUxYE', 'CQADAgAD2wYAAkn80Et8mVrPnEoyyBYE',
          'CQADAgADcQQAAuKn2EvxxDRi1xLq-xYE', 'CQADAgADhAQAAl_02UsxEX7mDgRL5RYE', 'CQADAgAD3AYAAkn80EvGTNAY4KErPBYE',
          'CQADAgAD3QYAAkn80EvQ13Zk_o8HaRYE', 'CQADAgADdQQAAuKn2EtDf6wHikeXvxYE', 'CQADAgADdgQAAuKn2EslNtgy_6fp0RYE',
          'CQADAgADeAQAAuKn2Etqv6ihz5LMXRYE', 'CQADAgADegQAAuKn2EtRt-NRDpb5XxYE', 'CQADAgADhgQAAl_02Ut1DFAy6WP96RYE',
          'CQADAgAD4AYAAkn80EusWEgzGgVX2xYE']

Sad_id =['CQADAgADrAQAAuKn2EvTYXbGCJhA_xYE', 'CQADAgADBAcAAkn80Eu8SpLIeqpjVRYE', 'CQADAgADsQQAAuKn2EtPdwj4_rVMVRYE',
         'CQADAgADBQcAAkn80Eu-KlWLkYBKQxYE', 'CQADAgADBwcAAkn80EvUDno6_qyRAxYE', 'CQADAgADtgQAAuKn2EtKyZd94LoUNBYE',
         'CQADAgADCQcAAkn80Eu3xJ4naMt1KxYE', 'CQADAgADuQQAAuKn2EseihnnQzqJlhYE', 'CQADAgADDAcAAkn80EvZBEsNzI-E2BYE',
         'CQADAgADvQQAAuKn2Esn6g1_-AzhpBYE', 'CQADAgADEAcAAkn80EvpdV9TvYllJhYE', 'CQADAgADEgcAAkn80EtjuwxR1GsjYBYE',
         'CQADAgADEwcAAkn80EtsP4Xs3euWYBYE', 'CQADAgADvwQAAuKn2Evre264AUZrVBYE', 'CQADAgADFQcAAkn80EtLifRoZpPlqBYE',
         'CQADAgADFgcAAkn80Es8hE7yupJumhYE', 'CQADAgADxAQAAuKn2EusuYR858FlOhYE', 'CQADAgADxQQAAuKn2Es9yEUAAbsZd3wWBA',
         'CQADAgADxgQAAuKn2EtgsAY-Qk2NTBYE', 'CQADAgADkwQAAl_02UtMpXeZxbwPJhYE', 'CQADAgADFwcAAkn80EsnFRJnwFEHIBYE']

Party_id = ['CQADAgADGgcAAkn80EunO4qC21oXhBYE', 'CQADAgADGwcAAkn80EtJ0tQioTcD1hYE', 'CQADAgADlQQAAl_02UtuDBWmwRopUxYE',
            'CQADAgADHAcAAkn80Eu9bwJVR4ONqhYE', 'CQADAgADHQcAAkn80EuAuKC_-C8PXBYE', 'CQADAgADHgcAAkn80EsjZJRFC9YiHhYE',
            'CQADAgADlgQAAl_02UtiTCZRYMOowBYE', 'CQADAgADHwcAAkn80Eu_dA0jrXvWyBYE', 'CQADAgADzQQAAuKn2EvGkvLqAvIzqhYE',
            'CQADAgADIAcAAkn80Es5oygtutY19hYE', 'CQADAgADIQcAAkn80Et6fzZARseZRBYE', 'CQADAgADIgcAAkn80EuaqluAR101-xYE',
            'CQADAgADzwQAAuKn2EtojAqCHL7HKRYE', 'CQADAgADJAcAAkn80EsYPQG2EJqNuRYE', 'CQADAgADJQcAAkn80EuttYM5UseC6BYE',
            'CQADAgAD0QQAAuKn2EuVThPRTk1sHBYE', 'CQADAgADKAcAAkn80EvZhTykoLc6KBYE', 'CQADAgADKgcAAkn80EuwpoemfskDhBYE',
            'CQADAgADLAcAAkn80EuEUDMIKx8qrRYE']

Nostalgia_id = ['CQADAgADNgcAAkn80EsMpozYaTayqxYE', 'CQADAgAD3AQAAuKn2Es9-Rq8QhvTwBYE', 'CQADAgAD3QQAAuKn2EtmjelNeBXGlxYE',
                'CQADAgAD3gQAAuKn2EszL3SVFwYDgxYE', 'CQADAgAD3wQAAuKn2EsUEep5yAFl6BYE', 'CQADAgADQAcAAkn80EvGkSNLFzkPlxYE',
                'CQADAgADQwcAAkn80Evh_Zac0CuMshYE', 'CQADAgAD4QQAAuKn2EujCjoU3xhyIxYE', 'CQADAgADRQcAAkn80Etl5AsSgTmwuhYE',
                'CQADAgADsAQAAl_02Usf3V1DPv82ExYE', 'CQADAgAD4gQAAuKn2EtsL8Y_WWDedxYE', 'CQADAgADRgcAAkn80Evd-ayXhFJyuhYE']

Studying_id =['CQADAgADSAcAAkn80Evzdad4kE0p0hYE', 'CQADAgADSQcAAkn80Eu7RP4eGL6DXxYE', 'CQADAgADtgQAAl_02Utkt0Mjl5QadxYE',
              'CQADAgADuQQAAl_02Uv78DwkidCO8RYE', 'CQADAgADSwcAAkn80EsZqsqFXp78VhYE', 'CQADAgADUAcAAkn80EtbweZDkY1wuxYE',
              'CQADAgADUgcAAkn80EuP7Yt9EriJQhYE', 'CQADAgADUwcAAkn80EsvkikOY98L7BYE', 'CQADAgADxAQAAl_02Uu34zFiPidIihYE',
              'CQADAgADVAcAAkn80Evb-JkvxNRGAxYE']

#Списки file_id аудио для жанров

Pop_id = ['CQADAgADVwcAAkn80EvnSJwmIolkghYE', 'CQADAgADWAcAAkn80EttsuWE73G6NxYE', 'CQADAgADWQcAAkn80Esaw8VDCNgQuBYE',
          'CQADAgAD5gQAAuKn2Es0_syZYRE_nRYE', 'CQADAgADWwcAAkn80EsEDFAbSm432xYE', 'CQADAgAD5wQAAuKn2EsCiw3vA-br0RYE',
          'CQADAgAD6AQAAuKn2Et3aVL--o7eqRYE', 'CQADAgADXQcAAkn80EtrdD_P_X274BYE', 'CQADAgADXgcAAkn80EvVM-VTHwgvQRYE',
          'CQADAgADXwcAAkn80Et2aXZMBintnBYE', 'CQADAgAD-gQAArqK2EsNeOWky-VmtxYE', 'CQADAgADYAcAAkn80EvpWKwyAgmOmRYE',
          'CQADAgADGAQAAuGH2EsJ9eW7Crr_UxYE', 'CQADAgADGQQAAuGH2EuRvX4Q7rec5BYE', 'CQADAgAD2AQAAl_02UuGfuUoqvNvtRYE',
          'CQADAgADGgQAAuGH2EtX58BlyBudpBYE']

Alternative_id = ['CQADAgADYwUAAuKn2Es-MiWHIwqmoBYE', 'CQADAgADGwUAAl_02Uth9nw0AW9qRxYE', 'CQADAgADRAQAAuGH2EuTp8kCersoqhYE',
                  'CQADAgADHgUAAl_02Uvuhw2XbVCuTBYE', 'CQADAgADRQQAAuGH2Eu4kys_Qe-vCBYE', 'CQADAgADRwQAAuGH2Et7FZgboBB5thYE',
                  'CQADAgADJQUAAl_02UvS-KPf-Knk1RYE', 'CQADAgADbgUAAuKn2EtEfsYkEOI2ThYE', 'CQADAgADTAQAAuGH2EuZTcx_EWMmFxYE',
                  'CQADAgADJwUAAl_02Ut9RxprO7VnaBYE', 'CQADAgADcAUAAuKn2Et1P683h7JkMRYE', 'CQADAgADTQQAAuGH2EtK09FwHW6I_hYE',
                  'CQADAgADTgQAAuGH2Eslw0CAW_QtsRYE']

Rock_id = ['CQADAgADUAQAAuGH2EvCd-q2s53z0xYE', 'CQADAgADcQUAAuKn2EuDmsUGmwNWORYE', 'CQADAgADKQUAAl_02UvI8AzTtv77wxYE',
           'CQADAgADdAUAAuKn2EvYjMpLAAGOW0oWBA', 'CQADAgADdQUAAuKn2EtQmLMnJvxp1hYE', 'CQADAgADUgQAAuGH2Eu1L6wm_pIrxRYE',
           'CQADAgADdwUAAuKn2EvY4IKsT0miKhYE', 'CQADAgADUwQAAuGH2EvOPGQ1cZODsBYE', 'CQADAgADeAUAAuKn2Et9FDj56wKmgBYE',
           'CQADAgADeQUAAuKn2EtPZDLO5syW7RYE', 'CQADAgADLwUAAl_02UsoJAkbMq-sqxYE', 'CQADAgADVQQAAuGH2Esd9sl00tq2vRYE',
           'CQADAgADfAUAAuKn2Eu3ff_BTfirRBYE', 'CQADAgADVgQAAuGH2Evr6_fyJ7mwHBYE', 'CQADAgADVwQAAuGH2EvfOYyySbbC7hYE',
           'CQADAgADfgUAAuKn2EuypwpO5ZGzxxYE', 'CQADAgADgAUAAuKn2Es-_Mh10DmlIRYE','CQADAgADWAQAAuGH2EvU4l208Pv_LBYE']

RandB_id = []

Electro_id = []

#Списки file_id аудио для сна

ForSleep_id = []

#Списки file_id для акустики

Acoustics_id = []

#new_list = random.sample(Christmas_id, k=5)
#def random_songs():
    #for i in range(len(new_list)):     В таком случае в ответе бота будут new_list[0] и тд
        #return new_list[i]


@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, 'Привет, как ты хочешь осуществить выбор музыки?', reply_markup=keyboard_menu)


@bot.message_handler(commands=['song'])
def find_file_ids(message):
    for file in os.listdir('/Users/elinaaptineeva/Desktop/rock/'):
        if file.split('.')[-1] == 'm4a' or 'mp3':
            songg = open('/Users/elinaaptineeva/Desktop/rock/'+file, 'rb')
            msg = bot.send_audio(message.chat.id, songg, None)
            bot.send_message(message.chat.id, msg.audio.file_id, reply_to_message_id=msg.message_id)
        time.sleep(50)


@bot.message_handler(func=lambda message: True, content_types=['text']) #Все в одной функции, так как все сообщения имеют один тип text
def send_text(message):
    if message.text.lower() == 'привет':
        bot.send_message(message.chat.id, 'Приветос')
    elif message.text.lower() == 'пока':
        bot.send_message(message.chat.id, 'Прощай......я буду скучать')
    elif message.text.lower() == 'назад':
        bot.send_message(message.chat.id, 'Как ты хочешь осуществить выбор музыки?', reply_markup=keyboard_menu)


        #Выбор для настроения
    elif message.text.lower() == 'по настроению':
        bot.send_message(message.chat.id, 'Какое у тебя настроение?', reply_markup=keyboard_mood)
    elif message.text.lower() == 'new year':
        bot.send_audio(message.chat.id, random.choice(NewYear_id), disable_notification=1) #не нашли функцию, позволяющюю отправлять несколько аудио в одном сообщении,
        bot.send_audio(message.chat.id, random.choice(NewYear_id), disable_notification=1) #поэтому приходится пять раз отправлять разные сообщения
        bot.send_audio(message.chat.id, random.choice(NewYear_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(NewYear_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(NewYear_id), disable_notification=1)
        bot.send_sticker(message.chat.id, 'CAADAgADEAADkp8eEQXW_GnuhWb_FgQ')

    elif message.text.lower() == 'in love':
        bot.send_audio(message.chat.id, random.choice(InLove_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(InLove_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(InLove_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(InLove_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(InLove_id), disable_notification=1)
        bot.send_sticker(message.chat.id, 'CAADAgADBAADkp8eEavc7j2rScUkFgQ')

    elif message.text.lower() == 'gym':
        bot.send_audio(message.chat.id, random.choice(Gym_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Gym_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Gym_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Gym_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Gym_id), disable_notification=1)
        bot.send_sticker(message.chat.id, 'CAADAgADXgADaUCAC_ik5ou43wUkFgQ')

    elif message.text.lower() == 'sad':
        bot.send_audio(message.chat.id, random.choice(Sad_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Sad_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Sad_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Sad_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Sad_id), disable_notification=1)
        bot.send_sticker(message.chat.id, 'CAADAQADoBIAApl_iALzktRqOg9uhhYE')

    elif message.text.lower() == 'party':
        bot.send_audio(message.chat.id, random.choice(Party_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Party_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Party_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Party_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Party_id), disable_notification=1)
        bot.send_sticker(message.chat.id, 'CAADAgADowADtvArFPC6O5csMnkcFgQ')

    elif message.text.lower() == 'nostalgia':
        bot.send_audio(message.chat.id, random.choice(Nostalgia_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Nostalgia_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Nostalgia_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Nostalgia_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Nostalgia_id), disable_notification=1)
        #bot.send_sticker(message.chat.id, '')

    elif message.text.lower() == 'studying':
        bot.send_audio(message.chat.id, random.choice(Studying_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Studying_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Studying_id), disable_notification=1)
        bot.send_sticker(message.chat.id, 'CAADAgADFAkAAlwCZQOykwORp3AchRYE')

    elif message.text.lower() == 'poetry':
        bot.send_audio(message.chat.id, 'CQADAgADLQcAAkn80Evi9KmmqvCqIhYE', disable_notification=1)
        bot.send_audio(message.chat.id, 'CQADAgAD2gQAAuKn2EvH5XFV329_sBYE', disable_notification=1)
        bot.send_audio(message.chat.id, 'CQADAgADLgcAAkn80EuRCVTz5CcrlRYE', disable_notification=1)
        bot.send_audio(message.chat.id, 'CQADAgADLwcAAkn80EvU12_SsfarvBYE', disable_notification=1)
        bot.send_sticker(message.chat.id, 'CAADAgADHQADijc4AAEw0RBgpCTPAAEWBA')


        #Выбор для жанра
    elif message.text.lower() == 'по жанрам':
        bot.send_message(message.chat.id, 'Какой ты хочешь жанр?', reply_markup=keyboard_genre)
    elif message.text.lower() == 'pop':
        bot.send_audio(message.chat.id, random.choice(Pop_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Pop_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Pop_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Pop_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Pop_id), disable_notification=1)
        #bot.send_sticker(message.chat.id, '')

    elif message.text.lower() == 'альтернатива':
        bot.send_audio(message.chat.id, random.choice(Alternative_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Alternative_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Alternative_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Alternative_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Alternative_id), disable_notification=1)
        #bot.send_sticker(message.chat.id, '')

    elif message.text.lower() == 'рок':
        bot.send_audio(message.chat.id, random.choice(Rock_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Rock_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Rock_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Rock_id), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(Rock_id), disable_notification=1)
        #bot.send_sticker(message.chat.id, '')

    elif message.text.lower() == 'r&b':
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        #bot.send_sticker(message.chat.id, '')

    elif message.text.lower() == 'электроника':
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        bot.send_audio(message.chat.id, random.choice(), disable_notification=1)
        #bot.send_sticker(message.chat.id, '')




    else:
        bot.send_sticker(message.chat.id, 'CAADAgADNAADkp8eEfIdTHZlMTXQFgQ')


bot.polling(none_stop = True, interval = 0)


